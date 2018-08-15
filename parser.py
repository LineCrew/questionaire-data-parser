import openpyxl
import json

raw = openpyxl.load_workbook('./800_01_0817.xlsx')

questionaire_result = []


print('start parsing data from excel file..')

print('sheet:', raw.sheetnames[0])
questionaires = []
sheet = raw[raw.sheetnames[0]]
for row in range(2, sheet.max_row + 1):
    if sheet['A' + str(row)].value is None:
        break
    questionaires.append({
        'number': int(
            str(sheet['A' + str(row)].value)
            + str(sheet['B' + str(row)].value)
            + str(sheet['C' + str(row)].value)),
        'content': sheet['D' + str(row)].value,
        'example': sheet['E' + str(row)].value if sheet['E' + str(row)].value is not None else '',
        'case1': sheet['F' + str(row)].value,
        'case2': sheet['G' + str(row)].value,
        'case3': sheet['H' + str(row)].value,
        'case4': sheet['I' + str(row)].value,
        'answer': sheet['J' + str(row)].value
    })

print('finishing data parse..')

output = json.dumps(questionaires, ensure_ascii=False)

print('writing parsed data to file..')

file = open('./output.json', 'w')
file.write(output)
file.close()

print('output succeeded to write in output.json file.')

# for sheet_name in raw.sheetnames:
